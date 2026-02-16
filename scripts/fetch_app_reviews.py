#!/usr/bin/env python3
"""Fetch Google Play and Apple App Store reviews from one input sheet."""

from __future__ import annotations

import argparse
import datetime as dt
import os
import sys
from collections import defaultdict
from pathlib import Path
from typing import Callable

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.review_collectors.apple_store import collect_apple_reviews
from src.review_collectors.common import (
    NormalizedReview,
    default_language_for_country,
    normalize_app_id,
    normalize_markets,
    parse_apple_app_id_from_link,
    parse_google_app_id_from_link,
    sanitize_filename,
    to_text,
)
from src.review_collectors.google_play import collect_google_reviews
from src.storage import date_token_to_iso, normalize_date_token, upsert_app_reviews


DEFAULT_MARKETS = ["KR", "US", "JP"]
DEFAULT_ROOT = Path(__file__).resolve().parents[1]

COLUMN_ALIASES = {
    "service_name": ["service_name", "앱 이름", "서비스명", "app_name"],
    "google_app_id": ["google_app_id", "google_id"],
    "apple_app_id": ["apple_app_id", "apple_id"],
    "google_link": ["google_link", "playstore", "play_store", "link"],
    "apple_link": ["apple_link", "appstore", "app_store", "link"],
    "markets": ["markets", "market", "국가", "대상국가"],
}

XLSX_COLUMNS = [
    "service_name",
    "store",
    "app_id",
    "country",
    "language",
    "rating",
    "title",
    "content",
    "reviewer_name",
    "review_created_at",
    "source_url",
]


def _normalize_row_keys(row: dict[object, object]) -> dict[str, object]:
    return {to_text(key).lower(): value for key, value in row.items()}


def _pick_value(row: dict[str, object], logical_name: str) -> str:
    aliases = COLUMN_ALIASES.get(logical_name, [])
    for alias in aliases:
        value = to_text(row.get(alias.lower()))
        if value:
            return value
    return ""


def normalize_target_row(row: dict[object, object], default_markets: list[str]) -> dict[str, object]:
    normalized_row = _normalize_row_keys(row)
    service_name = _pick_value(normalized_row, "service_name")
    google_app_id = normalize_app_id(_pick_value(normalized_row, "google_app_id"))
    apple_app_id = normalize_app_id(_pick_value(normalized_row, "apple_app_id"))
    google_link = _pick_value(normalized_row, "google_link")
    apple_link = _pick_value(normalized_row, "apple_link")

    if not google_app_id:
        google_app_id = parse_google_app_id_from_link(google_link)
    if not apple_app_id:
        apple_app_id = parse_apple_app_id_from_link(apple_link)

    markets_value = _pick_value(normalized_row, "markets")
    markets = normalize_markets(markets_value, default_markets)

    if not service_name:
        service_name = google_app_id or apple_app_id or "unnamed_service"

    return {
        "service_name": service_name,
        "google_app_id": google_app_id,
        "apple_app_id": apple_app_id,
        "markets": markets,
    }


def load_targets_from_excel(input_path: Path, default_markets: list[str]) -> list[dict[str, object]]:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("pandas is required to read input Excel files") from exc

    if not input_path.exists():
        raise FileNotFoundError(f"input file does not exist: {input_path}")

    frame = pd.read_excel(input_path)
    targets: list[dict[str, object]] = []
    for _, row in frame.iterrows():
        target = normalize_target_row(row.to_dict(), default_markets)
        if not target["google_app_id"] and not target["apple_app_id"]:
            continue
        targets.append(target)
    return targets


def _build_db_rows(
    *,
    timestamp: str,
    date_iso: str,
    service_name: str,
    store: str,
    app_id: str,
    country: str,
    language: str,
    reviews: list[NormalizedReview],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for review in reviews:
        rows.append(
            {
                "timestamp": timestamp,
                "date": date_iso,
                "service_name": service_name,
                "store": store,
                "app_id": app_id,
                "country": country,
                "language": language,
                "review_id": to_text(review["review_id"]),
                "review_created_at": to_text(review["review_created_at"]),
                "review_updated_at": to_text(review["review_updated_at"]),
                "rating": to_text(review["rating"]),
                "title": to_text(review["title"]),
                "content": to_text(review["content"]),
                "reviewer_name": to_text(review["reviewer_name"]),
                "source_url": to_text(review["source_url"]),
            }
        )
    return rows


def write_reviews_xlsx(path: Path, rows: list[dict[str, str]], sheet_name: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("openpyxl is required when --xlsx is enabled") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sanitize_filename(sheet_name)[:31]

    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10, color="000000")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    sheet.append(XLSX_COLUMNS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    for row in rows:
        sheet.append([row.get(column, "") for column in XLSX_COLUMNS])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = body_font
            cell.border = border
            if cell.column in {7, 8}:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 24,
        "B": 16,
        "C": 30,
        "D": 10,
        "E": 10,
        "F": 8,
        "G": 40,
        "H": 80,
        "I": 18,
        "J": 22,
        "K": 50,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def run_collection(
    *,
    root: Path,
    date_token: str,
    targets: list[dict[str, object]],
    limit_per_app_country: int,
    save_xlsx: bool,
    out_dir: Path,
    google_collector: Callable[[str, str, str, int], list[NormalizedReview]] = collect_google_reviews,
    apple_collector: Callable[[str, str, str, int], list[NormalizedReview]] = collect_apple_reviews,
) -> dict[str, object]:
    os.environ["KTRIPPEDIA_DATA_DIR"] = str(root / "data")
    date_iso = date_token_to_iso(normalize_date_token(date_token))

    summary: dict[str, object] = {
        "apps_total": len(targets),
        "apps_processed": 0,
        "rows_collected": 0,
        "rows_inserted": 0,
        "rows_updated": 0,
        "by_store_country": defaultdict(int),
        "errors": [],
    }

    for target in targets:
        service_name = to_text(target.get("service_name"))
        google_app_id = to_text(target.get("google_app_id"))
        apple_app_id = to_text(target.get("apple_app_id"))
        markets = [to_text(code).upper() for code in target.get("markets", []) if to_text(code)]
        if not markets:
            markets = DEFAULT_MARKETS[:]

        summary["apps_processed"] = int(summary["apps_processed"]) + 1
        for country in markets:
            language = default_language_for_country(country)

            def _save(store: str, app_id: str, reviews: list[NormalizedReview]) -> None:
                if not reviews:
                    return
                timestamp = dt.datetime.now().isoformat(timespec="seconds")
                db_rows = _build_db_rows(
                    timestamp=timestamp,
                    date_iso=date_iso,
                    service_name=service_name,
                    store=store,
                    app_id=app_id,
                    country=country,
                    language=language,
                    reviews=reviews,
                )
                stats = upsert_app_reviews(db_rows)
                summary["rows_collected"] = int(summary["rows_collected"]) + len(reviews)
                summary["rows_inserted"] = int(summary["rows_inserted"]) + int(stats["inserted"])
                summary["rows_updated"] = int(summary["rows_updated"]) + int(stats["updated"])
                key = f"{store}:{country}"
                summary["by_store_country"][key] += len(reviews)

                if save_xlsx:
                    output_name = f"{sanitize_filename(service_name)}_{store}_{app_id}_{country}.xlsx"
                    write_reviews_xlsx(out_dir / output_name, db_rows, sheet_name=f"{service_name}_{country}")

            if google_app_id:
                try:
                    google_reviews = google_collector(
                        google_app_id,
                        country,
                        language,
                        limit_per_app_country,
                    )
                    _save("google_play", google_app_id, google_reviews)
                except Exception as exc:
                    summary["errors"].append(f"google_play:{service_name}:{country}:{exc}")

            if apple_app_id:
                try:
                    apple_reviews = apple_collector(
                        apple_app_id,
                        country,
                        language,
                        limit_per_app_country,
                    )
                    _save("apple_app_store", apple_app_id, apple_reviews)
                except Exception as exc:
                    summary["errors"].append(f"apple_app_store:{service_name}:{country}:{exc}")

    summary["by_store_country"] = dict(summary["by_store_country"])
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fetch Google + Apple reviews in one run")
    parser.add_argument("--root", default=str(DEFAULT_ROOT), help="Project root")
    parser.add_argument("--input", default="data/app_review_targets.xlsx", help="Input Excel path")
    parser.add_argument("--date", default=dt.datetime.now().strftime("%Y%m%d"), help="Tracking date YYYYMMDD")
    parser.add_argument("--markets", default="KR,US,JP", help="Default markets, comma separated")
    parser.add_argument("--limit-per-app-country", type=int, default=0, help="0 means unlimited")
    parser.add_argument("--xlsx", action="store_true", help="Export per app/store/country xlsx files")
    parser.add_argument("--out-dir", default="data/reviews", help="XLSX output directory")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = Path(args.root).resolve()
    default_markets = normalize_markets(args.markets, DEFAULT_MARKETS)
    input_path = Path(args.input)
    if not input_path.is_absolute():
        input_path = root / input_path

    targets = load_targets_from_excel(input_path=input_path, default_markets=default_markets)
    summary = run_collection(
        root=root,
        date_token=args.date,
        targets=targets,
        limit_per_app_country=max(0, int(args.limit_per_app_country)),
        save_xlsx=bool(args.xlsx),
        out_dir=(root / args.out_dir) if not Path(args.out_dir).is_absolute() else Path(args.out_dir),
    )

    print(f"Targets loaded: {summary['apps_total']}")
    print(f"Targets processed: {summary['apps_processed']}")
    print(f"Rows collected: {summary['rows_collected']}")
    print(f"Rows inserted: {summary['rows_inserted']}")
    print(f"Rows updated: {summary['rows_updated']}")
    for key, count in sorted(dict(summary["by_store_country"]).items()):
        print(f"{key}: {count}")
    if summary["errors"]:
        print("Errors:")
        for item in summary["errors"]:
            print(f"- {item}")


if __name__ == "__main__":
    main()
